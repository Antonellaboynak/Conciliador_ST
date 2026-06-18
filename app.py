import streamlit as st
import pandas as pd
import numpy as np
from io import BytesIO
import re
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

st.set_page_config(page_title="Conciliación Hotelbeds - Aptour", layout="wide")

st.title("🏨 Conciliación Hotelbeds / Aptour")
st.markdown("Cargá los tres archivos para iniciar la conciliación.")

# ─── FILE UPLOAD ────────────────────────────────────────────────────────────
col1, col2, col3 = st.columns(3)
with col1:
    f_sistema = st.file_uploader("📄 listsistema (informe Hotelbeds Aptour)", type=["xlsx", "xls"])
with col2:
    f_operador = st.file_uploader("📄 listoperador (archivo del operador)", type=["xlsx", "xls"])
with col3:
    f_cobros = st.file_uploader("📄 listcobros (cobros y pagos Aptour)", type=["xlsx", "xls"])


def read_excel_any(file):
    """Read .xlsx or .xls using appropriate engine."""
    name = file.name.lower()
    if name.endswith(".xls"):
        return pd.read_excel(file, dtype=str, engine="xlrd")
    return pd.read_excel(file, dtype=str, engine="openpyxl")


def split_nom_cli(df):
    """Split NOM_CLI into APELLIDO (before first space) and NOMBRES (after first space)."""
    df = df.copy()
    nom = df["NOM_CLI"].fillna("")
    df["APELLIDO"] = nom.str.split(" ", n=1).str[0].str.strip()
    df["NOMBRES"] = nom.str.split(" ", n=1).str[1].fillna("").str.strip()
    return df


def normalize_res(value):
    """Remove all spaces from a reservation number string."""
    if pd.isna(value):
        return ""
    return re.sub(r"\s+", "", str(value)).strip()


def find_costous_col(df):
    """Find the COSTOUS column regardless of suffix noise like ',N,20,4'."""
    for col in df.columns:
        if col == "COSTOUS" or col.startswith("COSTOUS,"):
            return col
    return None


def build_detalle_map(df_sistema):
    """Build a dict: normalized_reservation_code -> (ID_RES, COSTOUS)."""
    costous_col = find_costous_col(df_sistema)
    mapping = {}
    for _, row in df_sistema.iterrows():
        id_res = normalize_res(row.get("ID_RES", ""))
        costous = row[costous_col] if costous_col else np.nan
        detalle = str(row.get("DETALLE", "") or "")
        parts = [p.strip() for p in detalle.split(",") if p.strip()]
        for p in parts:
            key = normalize_res(p)
            if key:
                mapping[key] = {"ID_RES": id_res, "COSTOUS": costous}
    return mapping


def match_by_name(apellido, nombres, cliente_str):
    """
    Return True if:
      - apellido matches any token in cliente_str  AND
      - at least one token of nombres matches any token in cliente_str.
    Case-insensitive.
    """
    if not apellido or not cliente_str:
        return False
    cli_tokens = set(t.upper() for t in str(cliente_str).split() if t.strip())
    ap_tokens = set(t.upper() for t in apellido.split() if t.strip())
    nom_tokens = set(t.upper() for t in nombres.split() if t.strip())

    ap_match = bool(ap_tokens & cli_tokens)
    nom_match = bool(nom_tokens & cli_tokens)
    return ap_match and nom_match


def build_cobros_name_map(df_cobros):
    """Build list of (APELLIDO, NOMBRES, ID_RES, VENTAUS, COBROSUS) from listcobros."""
    records = []
    for _, row in df_cobros.iterrows():
        records.append({
            "APELLIDO": str(row.get("APELLIDO", "") or "").upper(),
            "NOMBRES": str(row.get("NOMBRES", "") or "").upper(),
            "ID_RES": normalize_res(row.get("ID_RES", "")),
            "VENTAUS": row.get("VENTAUS", np.nan),
            "COBROSUS": row.get("COBROSUS", np.nan),
        })
    return records


def to_float(val):
    try:
        return float(str(val).replace(",", ".").strip())
    except Exception:
        return np.nan


def conciliate(df_sistema, df_operador, df_cobros):
    """Main conciliation logic. Returns df_operador with Estado filled and red_flags list."""

    # --- prep sistema ---
    df_s = df_sistema.copy()
    detalle_map = build_detalle_map(df_s)

    # --- prep cobros: split NOM_CLI ---
    df_c = split_nom_cli(df_cobros)
    cobros_records = build_cobros_name_map(df_c)

    # --- prep operador ---
    df_o = df_operador.copy()
    if "Estado" not in df_o.columns:
        df_o["Estado"] = np.nan

    red_flags = []  # row indices that need red highlight

    for idx, row in df_o.iterrows():
        # Skip if already has Estado
        if pd.notna(df_o.at[idx, "Estado"]) and str(df_o.at[idx, "Estado"]).strip():
            continue

        nro_res_raw = row.get("Nº de Reserva", "")
        nro_res = normalize_res(nro_res_raw)
        importe_op = to_float(row.get("Importe", np.nan))

        # ── PASS 1: match by Nº de Reserva vs DETALLE ──────────────────────
        matched_info = None
        if nro_res and nro_res in detalle_map:
            matched_info = detalle_map[nro_res]

        if matched_info:
            df_o.at[idx, "Estado"] = matched_info["ID_RES"]
            costous = to_float(matched_info["COSTOUS"])
            if pd.notna(costous) and pd.notna(importe_op):
                if abs(costous - importe_op) > 5:
                    red_flags.append(idx)
            continue  # done for this row

        # ── PASS 2: match by APELLIDO + NOMBRES vs Cliente ─────────────────
        cliente_str = str(row.get("Cliente", "") or "")
        for rec in cobros_records:
            if match_by_name(rec["APELLIDO"], rec["NOMBRES"], cliente_str):
                df_o.at[idx, "Estado"] = rec["ID_RES"]
                # Check monetary conditions
                ventaus = to_float(rec["VENTAUS"])
                cobrosus = to_float(rec["COBROSUS"])
                flag = False
                # Condition A: |COSTOUS - Importe| > 5
                # We don't have COSTOUS here from sistema directly,
                # so we look it up from detalle_map using ID_RES
                id_res_matched = rec["ID_RES"]
                costous_val = np.nan
                for info in detalle_map.values():
                    if info["ID_RES"] == id_res_matched:
                        costous_val = to_float(info["COSTOUS"])
                        break
                if pd.notna(costous_val) and pd.notna(importe_op):
                    if abs(costous_val - importe_op) > 5:
                        flag = True
                # Condition B: |VENTAUS - COBROSUS| > 500
                if pd.notna(ventaus) and pd.notna(cobrosus):
                    if abs(ventaus - cobrosus) > 500:
                        flag = True
                if flag:
                    red_flags.append(idx)
                break  # first name match wins

    return df_o, red_flags


def export_to_excel(df_operador, red_flags):
    """Export df_operador to Excel with red fills on flagged rows."""
    output = BytesIO()

    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        df_operador.to_excel(writer, index=False, sheet_name="Conciliación")
        wb = writer.book
        ws = writer.sheets["Conciliación"]

        # Style header
        header_fill = PatternFill("solid", fgColor="2F5496")
        header_font = Font(bold=True, color="FFFFFF", name="Arial", size=10)
        thin = Side(style="thin", color="BFBFBF")
        border = Border(left=thin, right=thin, top=thin, bottom=thin)

        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            cell.border = border

        ws.row_dimensions[1].height = 30

        # Find Estado column index
        estado_col = None
        for i, cell in enumerate(ws[1], 1):
            if cell.value == "Estado":
                estado_col = i
                break

        red_fill = PatternFill("solid", fgColor="FF0000")
        red_font = Font(bold=True, color="FFFFFF", name="Arial", size=10)
        normal_font = Font(name="Arial", size=10)
        alt_fill = PatternFill("solid", fgColor="EBF0FA")

        for row_num, row in enumerate(ws.iter_rows(min_row=2), start=2):
            df_row_idx = row_num - 2  # 0-based
            is_red = df_row_idx in red_flags
            for cell in row:
                cell.border = border
                if is_red and estado_col and cell.column == estado_col:
                    cell.fill = red_fill
                    cell.font = red_font
                elif row_num % 2 == 0:
                    cell.fill = alt_fill
                    cell.font = normal_font
                else:
                    cell.font = normal_font

        # Auto-width columns
        for col_cells in ws.columns:
            length = max(len(str(c.value or "")) for c in col_cells)
            ws.column_dimensions[get_column_letter(col_cells[0].column)].width = min(length + 4, 40)

    output.seek(0)
    return output


# ─── MAIN LOGIC ─────────────────────────────────────────────────────────────
if f_sistema and f_operador and f_cobros:
    with st.spinner("Procesando archivos..."):
        try:
            df_s = read_excel_any(f_sistema)
            df_o = read_excel_any(f_operador)
            df_c = read_excel_any(f_cobros)

            st.success(f"✅ Archivos cargados — Sistema: {len(df_s)} filas | Operador: {len(df_o)} filas | Cobros: {len(df_c)} filas")

            # Run conciliation
            df_result, red_flags = conciliate(df_s, df_o, df_c)

            # Stats
            total = len(df_result)
            matched = df_result["Estado"].notna() & (df_result["Estado"].astype(str).str.strip() != "")
            n_matched = matched.sum()
            n_red = len(red_flags)
            n_empty = total - n_matched

            st.markdown("---")
            m1, m2, m3, m4 = st.columns(4)
            m1.metric("Total filas operador", total)
            m2.metric("✅ Con coincidencia", int(n_matched))
            m3.metric("🔴 Diferencias marcadas", n_red)
            m4.metric("⬜ Sin coincidencia", int(n_empty))

            st.markdown("---")
            st.subheader("Vista previa del resultado")

            # Show with color hint
            def highlight_red(row):
                idx = row.name
                if idx in red_flags:
                    return ["background-color: #ffcccc"] * len(row)
                return [""] * len(row)

            st.dataframe(
                df_result.style.apply(highlight_red, axis=1),
                use_container_width=True,
                height=400,
            )

            # Export
            excel_bytes = export_to_excel(df_result, red_flags)

            st.download_button(
                label="⬇️ Descargar conciliación (.xlsx)",
                data=excel_bytes,
                file_name="conciliacion_hotelbeds.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )

        except Exception as e:
            st.error(f"❌ Error durante el procesamiento: {e}")
            st.exception(e)

else:
    st.info("👆 Cargá los tres archivos para comenzar.")

st.markdown("---")
st.caption("Aptour · Conciliación Hotelbeds · v1.0")
